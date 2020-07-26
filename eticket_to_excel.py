from openpyxl import *
from datetime import datetime

def write_data_to_excel():
    now = datetime.now()
    wb = load_workbook(r'C:\Users\Admin\Documents\Vaucer_za_npkrka.xlsx')
    ws = wb["Sheet1"]

    ticket_code = ["123456-Test"]
    entry = ["SKRADIN"]

    orig_voucher_num = ws.cell(12,9)
    new_voucher_num = (orig_voucher_num.value.split("/")[0])
    #print(new_voucher_num)
    mod_voucher_num = int(new_voucher_num) + 1
    final_voucher_num = (str(mod_voucher_num) + "/2020")
    #print(final_voucher_num_voucher_num)
    orig_voucher_num.value = final_voucher_num

    ticked_code_position = ws.cell(19,7)
    #print(ticked_code_position.value)
    ticked_code_position.value = str(ticket_code) #<----------FIX

    period_date_position = ws.cell(20,4)
    #print(period_date_position.value)
    period_date_position.value = now.strftime("%d.%m.%Y")
    #print(period_date_position.value)

    entry_position = ws.cell(18,1)
    print(entry_position.value)
    entry_position.value = ("ULAZ " + str(entry))
    print(entry_position.value)

    for i in entry_position.value:
        print(i)

    date_position = ws.cell(36,2)
    date_position.value = now.strftime("%d.%m.%Y")

    #wb.save(r'C:\Users\Admin\Documents\Vaucer_za_npkrka_copy.xlsx')
